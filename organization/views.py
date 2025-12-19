from datetime import timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import os
import re
from datetime import datetime

from accounts.forms import InvitationCreateForm
from accounts.models import Invitation, UserProfile
from news.models import News
from .forms import (
    CaldavSettingsForm,
    DepartmentForm,
    MailServerSettingsForm,
    MattermostSettingsForm,
    PortalSettingsForm,
    UnitForm,
)
from .models import (
    CaldavSettings,
    Department,
    DepartmentHead,
    MailServerSettings,
    MattermostSettings,
    PortalSettings,
    Unit,
    UnitHead,
    VacationPeriod,
    VacationRequest,
)
from .caldav_utils import get_user_meetings



def _is_portal_admin(user: User) -> bool:
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def home(request):
    """
    Главная страница портала: обзор в стиле 1С,
    с карточками основных разделов.
    Внизу отображается текст последней корпоративной новости (если она есть).
    """
    from datetime import datetime, date, timedelta
    from tasks.models import Task
    from django.contrib.auth.models import User
    
    latest_news = News.objects.prefetch_related("images").first()
    
    # Количество задач пользователя (не выполненных)
    tasks_count = 0
    tasks_today = []
    if request.user.is_authenticated:
        tasks_count = Task.objects.filter(created_by=request.user).exclude(status="done").count()
        # Задачи на сегодня (не выполненные)
        today = date.today()
        tasks_today = Task.objects.filter(
            created_by=request.user
        ).exclude(status="done").order_by("-created_at")[:10]  # Берем последние 10 задач
    
    # Все встречи пользователя
    all_meetings = []
    meetings_today_count = 0
    if request.user.is_authenticated:
        try:
            caldav_settings = CaldavSettings.get_solo()
            # Проверяем, настроен ли CalDAV сервер и настройки пользователя
            if caldav_settings.server_url:
                try:
                    profile = request.user.profile
                    if profile.caldav_email and profile.caldav_password:
                        today = date.today()
                        # Получаем встречи на ближайшие 30 дней
                        start_datetime = datetime.combine(today, datetime.min.time())
                        end_datetime = start_datetime + timedelta(days=30)
                        all_meetings = get_user_meetings(request.user, start_datetime, end_datetime)
                except AttributeError:
                    pass
                
                # Подсчитываем встречи на сегодня
                if all_meetings:
                    today_meetings = []
                    for meeting in all_meetings:
                        meeting_start = meeting.get("start")
                        if meeting_start:
                            if isinstance(meeting_start, datetime):
                                meeting_date = meeting_start.date()
                            elif isinstance(meeting_start, date):
                                meeting_date = meeting_start
                            else:
                                continue
                            if meeting_date == today:
                                today_meetings.append(meeting)
                    meetings_today_count = len(today_meetings)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error getting meetings: {str(e)}", exc_info=True)
            all_meetings = []
            meetings_today_count = 0
    
    # Последние два просмотренных сотрудника из сессии
    recent_employees = []
    if request.user.is_authenticated:
        viewed_employee_ids = request.session.get('viewed_employees', [])
        if viewed_employee_ids:
            # Берем последние 2 уникальных ID
            unique_ids = []
            for emp_id in reversed(viewed_employee_ids):
                if emp_id not in unique_ids:
                    unique_ids.append(emp_id)
                if len(unique_ids) >= 2:
                    break
            if unique_ids:
                recent_employees = User.objects.filter(
                    id__in=unique_ids,
                    is_active=True
                ).select_related('profile').order_by('-id')[:2]
    
    # Последние новости
    recent_news = News.objects.all().order_by('-created_at')[:5]
    
    # Сегодняшняя дата для шаблона
    today = date.today()
    
    # Получаем порядок блоков из профиля пользователя
    blocks_order = []
    if request.user.is_authenticated:
        try:
            from django.db import connection
            # Проверяем существование колонки через SQL
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'accounts_userprofile' 
                    AND column_name = 'dashboard_blocks_order'
                """)
                column_exists = cursor.fetchone() is not None
            
            if column_exists:
                from accounts.models import UserProfile
                profile = UserProfile.objects.get(user=request.user)
                blocks_order = profile.dashboard_blocks_order or []
        except Exception:
            blocks_order = []
    
    return render(
        request,
        "home.html",
        {
            "latest_news": latest_news,
            "tasks_count": tasks_count,
            "tasks_today": tasks_today,
            "meetings_today_count": meetings_today_count,
            "all_meetings": all_meetings,
            "recent_employees": recent_employees,
            "recent_news": recent_news,
            "today": today,
            "blocks_order": blocks_order,
        },
    )


def tasks(request):
    # делегируем вьюху «Мои задачи» из приложения tasks
    from tasks.views import my_tasks

    return my_tasks(request)


@login_required
def meetings(request):
    """
    Раздел «Мои встречи» — отображение встреч из CalDAV календаря.
    Использует настройки пользователя (caldav_email и caldav_password) из профиля.
    URL календаря формируется по шаблону: $server/SOGo/dav/$user_email/Calendar/personal/
    """
    from datetime import datetime, timedelta
    
    meetings_list = []
    error_message = None
    caldav_configured = False
    user_caldav_configured = False
    
    # Проверяем, настроен ли CalDAV сервер в системных настройках
    caldav_settings = CaldavSettings.get_solo()
    if caldav_settings.server_url:
        caldav_configured = True
        
        # Проверяем, настроены ли настройки пользователя
        try:
            profile = request.user.profile
            if profile.caldav_email and profile.caldav_password:
                user_caldav_configured = True
        except AttributeError:
            pass
        
        # Получаем встречи пользователя, если настройки пользователя указаны
        if user_caldav_configured:
            try:
                # Получаем встречи на ближайшие 30 дней
                start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = start_date + timedelta(days=30)
                
                meetings_list = get_user_meetings(request.user, start_date, end_date)
            except Exception as e:
                error_message = f"Ошибка при получении встреч из CalDAV: {str(e)}"
                import logging
                logger = logging.getLogger(__name__)
                logger.error(error_message, exc_info=True)
    
    return render(
        request,
        "organization/meetings.html",
        {
            "meetings": meetings_list,
            "error_message": error_message,
            "caldav_configured": caldav_configured,
            "user_caldav_configured": user_caldav_configured,
        },
    )


def employees(request):
    """
    Каталог сотрудников компании.

    Отображает всех активных пользователей портала вместе с их профилями
    и поддерживает простой поиск по ФИО, email, телефону, департаменту и отделу.
    """

    query = request.GET.get("q", "").strip()
    employee_id = request.GET.get("id", "").strip()

    # Если передан ID сотрудника, сохраняем его в сессию для "Последние контакты"
    if employee_id and request.user.is_authenticated:
        try:
            emp_id = int(employee_id)
            viewed_employees = request.session.get('viewed_employees', [])
            # Добавляем в начало списка, если его там еще нет
            if emp_id not in viewed_employees:
                viewed_employees.insert(0, emp_id)
            else:
                # Перемещаем в начало, если уже есть
                viewed_employees.remove(emp_id)
                viewed_employees.insert(0, emp_id)
            # Ограничиваем список последними 10 просмотренными
            viewed_employees = viewed_employees[:10]
            request.session['viewed_employees'] = viewed_employees
        except (ValueError, TypeError):
            pass

    employees_qs = (
        User.objects.select_related("profile")
        .filter(is_active=True)
        .order_by("username")
    )

    if query:
        employees_qs = employees_qs.filter(
            Q(username__icontains=query)
            | Q(email__icontains=query)
            | Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(profile__first_name__icontains=query)
            | Q(profile__last_name__icontains=query)
            | Q(profile__middle_name__icontains=query)
            | Q(profile__department__name__icontains=query)
            | Q(profile__unit__name__icontains=query)
            | Q(profile__position__name__icontains=query)
            | Q(profile__phone_internal__icontains=query)
        )

    return render(
        request,
        "organization/employees.html",
        {
            "title": "Сотрудники",
            "subtitle": "Каталог сотрудников компании с поиском по ФИО, email, отделу и телефону.",
            "employees": employees_qs,
            "query": query,
        },
    )


def org_chart(request):
    """
    Визуализация организационной структуры компании.
    Показывает все департаменты, отделы и сотрудников в виде схемы.
    Учитывает иерархию департаментов.
    """
    # Получаем только корневые департаменты (без родительских)
    # Используем defer для исключения новых полей до применения миграции
    root_departments = (
        Department.objects.filter(parent_department__isnull=True)
        .defer("department_head")
        .select_related("head", "parent_department")
        .prefetch_related("units__head", "units__employees__profile", "sub_departments")
        .order_by("name")
    )
    
    # Все департаменты для построения полной иерархии
    all_departments = (
        Department.objects.defer("department_head")
        .select_related("head", "parent_department")
        .prefetch_related("units__head", "units__employees__profile", "sub_departments")
        .all()
    )
    
    # Отделы, которые не принадлежат ни одному департаменту
    standalone_units = Unit.objects.filter(departments__isnull=True).defer("unit_head").select_related("head").prefetch_related("employees__profile").order_by("name")
    
    return render(
        request,
        "organization/org_chart.html",
        {
            "root_departments": root_departments,
            "all_departments": all_departments,
            "standalone_units": standalone_units,
        },
    )


@login_required
@user_passes_test(_is_portal_admin)
def units(request):
    """
    Раздел «Отделы и департаменты».
    Позволяет:
    - создавать департаменты и назначать руководителя;
    - создавать отделы, назначать руководителя и привязывать к департаменту;
    - добавлять сотрудников в отделы.
    Доступно только администраторам.
    """
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        if form_type == "department":
            # Создаем копию POST данных для возможной модификации
            post_data = request.POST.copy()
            
            # Проверяем, не добавлен ли новый руководитель из списка сотрудников
            new_head_user_id = request.POST.get("new_department_head_user")
            if new_head_user_id:
                try:
                    new_head_user = User.objects.get(pk=new_head_user_id, is_active=True)
                    department_name = request.POST.get("name", "")
                    if department_name:
                        # Создаем или получаем DepartmentHead для этого сотрудника и департамента
                        department_head, created = DepartmentHead.objects.get_or_create(
                            user=new_head_user,
                            department_name=department_name,
                            defaults={}
                        )
                        # Обновляем данные из UserProfile
                        if hasattr(new_head_user, "profile"):
                            profile = new_head_user.profile
                            department_head.last_name = profile.last_name or ""
                            department_head.first_name = profile.first_name or ""
                            department_head.middle_name = profile.middle_name or ""
                            department_head.department = str(profile.department) if profile.department else ""
                            department_head.unit = str(profile.unit) if profile.unit else ""
                            department_head.position = str(profile.position) if profile.position else ""
                            department_head.phone_personal = profile.phone_personal or ""
                            department_head.phone_internal = profile.phone_internal or ""
                            department_head.photo = profile.photo
                        department_head.save()
                        # Устанавливаем department_head в POST данных
                        post_data['dept-department_head'] = str(department_head.pk)
                except User.DoesNotExist:
                    pass
            
            department_form = DepartmentForm(post_data, prefix="dept")
            unit_form = UnitForm(prefix="unit")
            if department_form.is_valid():
                department = department_form.save(commit=False)
                # Синхронизируем head с department_head.user
                if department.department_head:
                    department.head = department.department_head.user
                    # Обновляем department_name в DepartmentHead, если изменилось название департамента
                    department.department_head.department_name = department.name
                    # Обновляем данные сотрудника из UserProfile
                    if hasattr(department.department_head.user, "profile"):
                        profile = department.department_head.user.profile
                        department.department_head.last_name = profile.last_name or ""
                        department.department_head.first_name = profile.first_name or ""
                        department.department_head.middle_name = profile.middle_name or ""
                        department.department_head.department = str(profile.department) if profile.department else ""
                        department.department_head.unit = str(profile.unit) if profile.unit else ""
                        department.department_head.position = str(profile.position) if profile.position else ""
                        department.department_head.phone_personal = profile.phone_personal or ""
                        department.department_head.phone_internal = profile.phone_internal or ""
                        department.department_head.photo = profile.photo
                    department.department_head.save()
                department.save()
                department_form.save_m2m()
                # Автоматически добавляем руководителя департамента во все отделы департамента
                if department.head:
                    for unit in department.units.all():
                        unit.employees.add(department.head)
                    # Также добавляем руководителя во все отделы подчинённых департаментов
                    def add_head_to_sub_departments(dept):
                        for sub_dept in dept.sub_departments.all():
                            for unit in sub_dept.units.all():
                                unit.employees.add(dept.head)
                            add_head_to_sub_departments(sub_dept)
                    add_head_to_sub_departments(department)
                return redirect("units")
        elif form_type == "unit":
            # Создаем копию POST данных для возможной модификации
            post_data = request.POST.copy()
            
            # Проверяем, не добавлен ли новый руководитель из списка сотрудников
            new_head_user_id = request.POST.get("new_unit_head_user")
            if new_head_user_id:
                try:
                    new_head_user = User.objects.get(pk=new_head_user_id, is_active=True)
                    unit_name = request.POST.get("name", "")
                    if unit_name:
                        # Создаем или получаем UnitHead для этого сотрудника и отдела
                        unit_head, created = UnitHead.objects.get_or_create(
                            user=new_head_user,
                            unit_name=unit_name,
                            defaults={}
                        )
                        # Обновляем данные из UserProfile
                        if hasattr(new_head_user, "profile"):
                            profile = new_head_user.profile
                            unit_head.last_name = profile.last_name or ""
                            unit_head.first_name = profile.first_name or ""
                            unit_head.middle_name = profile.middle_name or ""
                            unit_head.department = str(profile.department) if profile.department else ""
                            unit_head.unit = str(profile.unit) if profile.unit else ""
                            unit_head.position = str(profile.position) if profile.position else ""
                            unit_head.phone_personal = profile.phone_personal or ""
                            unit_head.phone_internal = profile.phone_internal or ""
                            unit_head.photo = profile.photo
                        unit_head.save()
                        # Устанавливаем unit_head в POST данных
                        post_data['unit-unit_head'] = str(unit_head.pk)
                except User.DoesNotExist:
                    pass
            
            department_form = DepartmentForm(prefix="dept")
            unit_form = UnitForm(post_data, prefix="unit")
            if unit_form.is_valid():
                unit = unit_form.save(commit=False)
                # Синхронизируем head с unit_head.user
                if unit.unit_head:
                    unit.head = unit.unit_head.user
                    # Обновляем unit_name в UnitHead, если изменилось название отдела
                    unit.unit_head.unit_name = unit.name
                    # Обновляем данные сотрудника из UserProfile
                    if hasattr(unit.unit_head.user, "profile"):
                        profile = unit.unit_head.user.profile
                        unit.unit_head.last_name = profile.last_name or ""
                        unit.unit_head.first_name = profile.first_name or ""
                        unit.unit_head.middle_name = profile.middle_name or ""
                        unit.unit_head.department = str(profile.department) if profile.department else ""
                        unit.unit_head.unit = str(profile.unit) if profile.unit else ""
                        unit.unit_head.position = str(profile.position) if profile.position else ""
                        unit.unit_head.phone_personal = profile.phone_personal or ""
                        unit.unit_head.phone_internal = profile.phone_internal or ""
                        unit.unit_head.photo = profile.photo
                    unit.unit_head.save()
                unit.save()
                unit_form.save_m2m()
                # Автоматически добавляем руководителя отдела в сотрудники отдела
                if unit.head:
                    unit.employees.add(unit.head)
                    # Также добавляем руководителя отдела во все отделы департаментов, к которым относится отдел
                    for department in unit.departments.all():
                        for dept_unit in department.units.all():
                            dept_unit.employees.add(unit.head)
                        # Добавляем руководителя отдела во все отделы родительских департаментов
                        parent = department.parent_department
                        while parent:
                            for parent_unit in parent.units.all():
                                parent_unit.employees.add(unit.head)
                            parent = parent.parent_department
                return redirect("units")
        elif form_type == "add_department_head":
            # Добавление нового руководителя департамента из списка сотрудников
            user_id = request.POST.get("user_id")
            department_name = request.POST.get("department_name", "")
            if user_id and department_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    department_head, created = DepartmentHead.objects.get_or_create(
                        user=user,
                        department_name=department_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        department_head.last_name = profile.last_name or ""
                        department_head.first_name = profile.first_name or ""
                        department_head.middle_name = profile.middle_name or ""
                        department_head.department = str(profile.department) if profile.department else ""
                        department_head.unit = str(profile.unit) if profile.unit else ""
                        department_head.position = str(profile.position) if profile.position else ""
                        department_head.phone_personal = profile.phone_personal or ""
                        department_head.phone_internal = profile.phone_internal or ""
                        department_head.photo = profile.photo
                    department_head.save()
                except User.DoesNotExist:
                    pass
            return redirect("units")
        elif form_type == "remove_department_head":
            # Удаление руководителя департамента
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    DepartmentHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass
            return redirect("units")
        elif form_type == "add_unit_head":
            # Добавление нового руководителя отдела из списка сотрудников
            user_id = request.POST.get("user_id")
            unit_name = request.POST.get("unit_name", "")
            if user_id and unit_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    unit_head, created = UnitHead.objects.get_or_create(
                        user=user,
                        unit_name=unit_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        unit_head.last_name = profile.last_name or ""
                        unit_head.first_name = profile.first_name or ""
                        unit_head.middle_name = profile.middle_name or ""
                        unit_head.department = str(profile.department) if profile.department else ""
                        unit_head.unit = str(profile.unit) if profile.unit else ""
                        unit_head.position = str(profile.position) if profile.position else ""
                        unit_head.phone_personal = profile.phone_personal or ""
                        unit_head.phone_internal = profile.phone_internal or ""
                        unit_head.photo = profile.photo
                    unit_head.save()
                except User.DoesNotExist:
                    pass
            return redirect("units")
        elif form_type == "remove_unit_head":
            # Удаление руководителя отдела
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    UnitHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass
            return redirect("units")
        else:
            department_form = DepartmentForm(prefix="dept")
            unit_form = UnitForm(prefix="unit")
    else:
        department_form = DepartmentForm(prefix="dept")
        unit_form = UnitForm(prefix="unit")

    # Используем defer для исключения новых полей до применения миграции
    departments = (
        Department.objects.defer("department_head")
        .select_related("head", "parent_department")
        .prefetch_related("units__head", "units__employees", "sub_departments")
    )
    units_qs = Unit.objects.defer("unit_head").select_related("head").prefetch_related("employees", "departments")
    
    # Получаем всех активных сотрудников для выбора
    all_employees = User.objects.filter(is_active=True).select_related('profile').order_by('username')
    
    # Получаем списки руководителей
    department_heads = DepartmentHead.objects.select_related('user').order_by('department_name', 'last_name', 'first_name')
    unit_heads = UnitHead.objects.select_related('user').order_by('unit_name', 'last_name', 'first_name')

    return render(
        request,
        "organization/units.html",
        {
            "department_form": department_form,
            "unit_form": unit_form,
            "departments": departments,
            "units": units_qs,
            "all_employees": all_employees,
            "department_heads": department_heads,
            "unit_heads": unit_heads,
        },
    )


@login_required
@user_passes_test(_is_portal_admin)
def unit_edit(request, pk):
    """
    Редактирование отдела: изменение названия, руководителя, добавление/исключение сотрудников.
    """
    unit = get_object_or_404(Unit, pk=pk)
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "update":
            # Создаем копию POST данных для возможной модификации
            post_data = request.POST.copy()
            
            # Проверяем, не добавлен ли новый руководитель из списка сотрудников
            new_head_user_id = request.POST.get("new_unit_head_user")
            if new_head_user_id:
                try:
                    new_head_user = User.objects.get(pk=new_head_user_id, is_active=True)
                    unit_name = request.POST.get("name", unit.name)
                    if unit_name:
                        # Создаем или получаем UnitHead для этого сотрудника и отдела
                        unit_head, created = UnitHead.objects.get_or_create(
                            user=new_head_user,
                            unit_name=unit_name,
                            defaults={}
                        )
                        # Обновляем данные из UserProfile
                        if hasattr(new_head_user, "profile"):
                            profile = new_head_user.profile
                            unit_head.last_name = profile.last_name or ""
                            unit_head.first_name = profile.first_name or ""
                            unit_head.middle_name = profile.middle_name or ""
                            unit_head.department = str(profile.department) if profile.department else ""
                            unit_head.unit = str(profile.unit) if profile.unit else ""
                            unit_head.position = str(profile.position) if profile.position else ""
                            unit_head.phone_personal = profile.phone_personal or ""
                            unit_head.phone_internal = profile.phone_internal or ""
                            unit_head.photo = profile.photo
                        unit_head.save()
                        # Устанавливаем unit_head в POST данных
                        post_data['unit-unit_head'] = str(unit_head.pk)
                except User.DoesNotExist:
                    pass
            
            form = UnitForm(post_data, instance=unit, prefix="unit")
            if form.is_valid():
                unit = form.save(commit=False)
                # Синхронизируем head с unit_head.user
                if unit.unit_head:
                    unit.head = unit.unit_head.user
                    # Обновляем unit_name в UnitHead, если изменилось название отдела
                    unit.unit_head.unit_name = unit.name
                    # Обновляем данные сотрудника из UserProfile
                    if hasattr(unit.unit_head.user, "profile"):
                        profile = unit.unit_head.user.profile
                        unit.unit_head.last_name = profile.last_name or ""
                        unit.unit_head.first_name = profile.first_name or ""
                        unit.unit_head.middle_name = profile.middle_name or ""
                        unit.unit_head.department = profile.department or ""
                        unit.unit_head.unit = profile.unit or ""
                        unit.unit_head.position = profile.position or ""
                        unit.unit_head.phone_personal = profile.phone_personal or ""
                        unit.unit_head.phone_internal = profile.phone_internal or ""
                        unit.unit_head.photo = profile.photo
                    unit.unit_head.save()
                unit.save()
                form.save_m2m()
                # Автоматически добавляем руководителя отдела в сотрудники отдела
                if unit.head:
                    unit.employees.add(unit.head)
                    # Также добавляем руководителя отдела во все отделы департаментов, к которым относится отдел
                    for department in unit.departments.all():
                        for dept_unit in department.units.all():
                            dept_unit.employees.add(unit.head)
                        # Добавляем руководителя отдела во все отделы родительских департаментов
                        parent = department.parent_department
                        while parent:
                            for parent_unit in parent.units.all():
                                parent_unit.employees.add(unit.head)
                            parent = parent.parent_department
                return redirect("units")
        elif action == "remove_employee":
            employee_id = request.POST.get("employee_id")
            if employee_id:
                try:
                    employee = User.objects.get(pk=employee_id)
                    unit.employees.remove(employee)
                    return redirect("unit_edit", pk=unit.pk)
                except User.DoesNotExist:
                    pass
        elif action == "add_employee":
            employee_id = request.POST.get("employee_id")
            if employee_id:
                try:
                    employee = User.objects.get(pk=employee_id)
                    unit.employees.add(employee)
                    return redirect("unit_edit", pk=unit.pk)
                except User.DoesNotExist:
                    pass
        elif action == "add_unit_head":
            # Добавление нового руководителя отдела
            user_id = request.POST.get("user_id")
            unit_name = request.POST.get("unit_name", unit.name)
            if user_id and unit_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    unit_head, created = UnitHead.objects.get_or_create(
                        user=user,
                        unit_name=unit_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        unit_head.last_name = profile.last_name or ""
                        unit_head.first_name = profile.first_name or ""
                        unit_head.middle_name = profile.middle_name or ""
                        unit_head.department = str(profile.department) if profile.department else ""
                        unit_head.unit = str(profile.unit) if profile.unit else ""
                        unit_head.position = str(profile.position) if profile.position else ""
                        unit_head.phone_personal = profile.phone_personal or ""
                        unit_head.phone_internal = profile.phone_internal or ""
                        unit_head.photo = profile.photo
                    unit_head.save()
                except User.DoesNotExist:
                    pass
            return redirect("unit_edit", pk=unit.pk)
        elif action == "remove_unit_head":
            # Удаление руководителя отдела
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    UnitHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass
            return redirect("unit_edit", pk=unit.pk)
    else:
        form = UnitForm(instance=unit, prefix="unit")
    
    # Получаем всех активных пользователей, которые ещё не в этом отделе
    existing_employee_ids = unit.employees.values_list("id", flat=True)
    available_employees = User.objects.filter(is_active=True).exclude(id__in=existing_employee_ids).order_by("username")
    
    # Получаем всех активных сотрудников для выбора руководителя
    all_employees = User.objects.filter(is_active=True).select_related('profile').order_by('username')
    
    # Получаем список руководителей отделов
    unit_heads = UnitHead.objects.select_related('user').order_by('unit_name', 'last_name', 'first_name')
    
    return render(
        request,
        "organization/unit_edit.html",
        {
            "unit": unit,
            "form": form,
            "available_employees": available_employees,
            "all_employees": all_employees,
            "unit_heads": unit_heads,
        },
    )


@login_required
@user_passes_test(_is_portal_admin)
def department_edit(request, pk):
    """
    Редактирование департамента: изменение названия, руководителя, добавление/исключение отделов.
    """
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "update":
            # Создаем копию POST данных для возможной модификации
            post_data = request.POST.copy()
            
            # Проверяем, не добавлен ли новый руководитель из списка сотрудников
            new_head_user_id = request.POST.get("new_department_head_user")
            if new_head_user_id:
                try:
                    new_head_user = User.objects.get(pk=new_head_user_id, is_active=True)
                    department_name = request.POST.get("name", department.name)
                    if department_name:
                        # Создаем или получаем DepartmentHead для этого сотрудника и департамента
                        department_head, created = DepartmentHead.objects.get_or_create(
                            user=new_head_user,
                            department_name=department_name,
                            defaults={}
                        )
                        # Обновляем данные из UserProfile
                        if hasattr(new_head_user, "profile"):
                            profile = new_head_user.profile
                            department_head.last_name = profile.last_name or ""
                            department_head.first_name = profile.first_name or ""
                            department_head.middle_name = profile.middle_name or ""
                            department_head.department = str(profile.department) if profile.department else ""
                            department_head.unit = str(profile.unit) if profile.unit else ""
                            department_head.position = str(profile.position) if profile.position else ""
                            department_head.phone_personal = profile.phone_personal or ""
                            department_head.phone_internal = profile.phone_internal or ""
                            department_head.photo = profile.photo
                        department_head.save()
                        # Устанавливаем department_head в POST данных
                        post_data['dept-department_head'] = str(department_head.pk)
                except User.DoesNotExist:
                    pass
            
            form = DepartmentForm(post_data, instance=department, prefix="dept")
            if form.is_valid():
                department = form.save(commit=False)
                # Синхронизируем head с department_head.user
                if department.department_head:
                    department.head = department.department_head.user
                    # Обновляем department_name в DepartmentHead, если изменилось название департамента
                    department.department_head.department_name = department.name
                    # Обновляем данные сотрудника из UserProfile
                    if hasattr(department.department_head.user, "profile"):
                        profile = department.department_head.user.profile
                        department.department_head.last_name = profile.last_name or ""
                        department.department_head.first_name = profile.first_name or ""
                        department.department_head.middle_name = profile.middle_name or ""
                        department.department_head.department = str(profile.department) if profile.department else ""
                        department.department_head.unit = str(profile.unit) if profile.unit else ""
                        department.department_head.position = str(profile.position) if profile.position else ""
                        department.department_head.phone_personal = profile.phone_personal or ""
                        department.department_head.phone_internal = profile.phone_internal or ""
                        department.department_head.photo = profile.photo
                    department.department_head.save()
                department.save()
                form.save_m2m()
                # Автоматически добавляем руководителя департамента во все отделы департамента
                if department.head:
                    for unit in department.units.all():
                        unit.employees.add(department.head)
                    # Также добавляем руководителя во все отделы подчинённых департаментов
                    def add_head_to_sub_departments(dept):
                        for sub_dept in dept.sub_departments.all():
                            for unit in sub_dept.units.all():
                                unit.employees.add(dept.head)
                            add_head_to_sub_departments(sub_dept)
                    add_head_to_sub_departments(department)
                return redirect("units")
        elif action == "add_unit":
            unit_id = request.POST.get("unit_id")
            if unit_id:
                try:
                    unit = Unit.objects.get(pk=unit_id)
                    department.units.add(unit)
                    # Автоматически добавляем руководителя департамента в новый отдел
                    if department.head:
                        unit.employees.add(department.head)
                    # Автоматически добавляем руководителя отдела во все отделы департамента
                    if unit.head:
                        for dept_unit in department.units.all():
                            dept_unit.employees.add(unit.head)
                    # Также добавляем руководителей всех родительских департаментов в новый отдел
                    parent = department.parent_department
                    while parent:
                        if parent.head:
                            unit.employees.add(parent.head)
                        parent = parent.parent_department
                    return redirect("department_edit", pk=department.pk)
                except Unit.DoesNotExist:
                    pass
        elif action == "remove_unit":
            unit_id = request.POST.get("unit_id")
            if unit_id:
                try:
                    unit = Unit.objects.get(pk=unit_id)
                    department.units.remove(unit)
                    return redirect("department_edit", pk=department.pk)
                except Unit.DoesNotExist:
                    pass
        elif action == "add_department_head":
            # Добавление нового руководителя департамента
            user_id = request.POST.get("user_id")
            department_name = request.POST.get("department_name", department.name)
            if user_id and department_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    department_head, created = DepartmentHead.objects.get_or_create(
                        user=user,
                        department_name=department_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        department_head.last_name = profile.last_name or ""
                        department_head.first_name = profile.first_name or ""
                        department_head.middle_name = profile.middle_name or ""
                        department_head.department = str(profile.department) if profile.department else ""
                        department_head.unit = str(profile.unit) if profile.unit else ""
                        department_head.position = str(profile.position) if profile.position else ""
                        department_head.phone_personal = profile.phone_personal or ""
                        department_head.phone_internal = profile.phone_internal or ""
                        department_head.photo = profile.photo
                    department_head.save()
                except User.DoesNotExist:
                    pass
            return redirect("department_edit", pk=department.pk)
        elif action == "remove_department_head":
            # Удаление руководителя департамента
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    DepartmentHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass
            return redirect("department_edit", pk=department.pk)
    else:
        form = DepartmentForm(instance=department, prefix="dept")
    
    # Получаем все отделы, которые ещё не в этом департаменте
    existing_unit_ids = department.units.values_list("id", flat=True)
    available_units = Unit.objects.exclude(id__in=existing_unit_ids).order_by("name")
    
    # Получаем все департаменты, которые могут быть родительскими (исключаем текущий и его поддепартаменты)
    exclude_ids = [department.pk]
    # Рекурсивно собираем все поддепартаменты, чтобы исключить их из списка возможных родителей
    def get_sub_department_ids(dept):
        ids = [dept.pk]
        for sub_dept in dept.sub_departments.all():
            ids.extend(get_sub_department_ids(sub_dept))
        return ids
    
    exclude_ids.extend(get_sub_department_ids(department))
    available_parents = Department.objects.exclude(id__in=exclude_ids).order_by("name")
    
    # Получаем всех активных сотрудников для выбора руководителя
    all_employees = User.objects.filter(is_active=True).select_related('profile').order_by('username')
    
    # Получаем список руководителей департаментов
    department_heads = DepartmentHead.objects.select_related('user').order_by('department_name', 'last_name', 'first_name')
    
    return render(
        request,
        "organization/department_edit.html",
        {
            "department": department,
            "form": form,
            "available_units": available_units,
            "available_parents": available_parents,
            "all_employees": all_employees,
            "department_heads": department_heads,
        },
    )


@login_required
def chats(request):
    """
    Раздел чатов с интеграцией Mattermost.
    """
    from .mattermost_client import MattermostClient
    from .models import MattermostSettings

    error_message = None
    channels = []
    regular_channels = []
    direct_channels = []
    all_users = []
    teams = []
    selected_channel = None
    channel_posts = []
    channel_members = []
    mattermost_configured = False
    user_configured = False
    login_success = False

    # Проверяем настройки сервера
    try:
        mm_settings = MattermostSettings.get_solo()
        mattermost_configured = bool(mm_settings.server_url)
    except Exception:
        pass

    # Проверяем настройки пользователя
    try:
        profile = request.user.profile
        user_configured = bool(profile.mattermost_username and profile.mattermost_password)
    except Exception:
        profile = None

    # Создаем директорию для файлов Mattermost, если её нет
    try:
        os.makedirs(settings.MATTERMOST_FILES_ROOT, exist_ok=True)
    except Exception:
        pass
    
    # Если пользователь настроен, получаем данные из Mattermost
    if user_configured and mattermost_configured and profile:
        try:
            client = MattermostClient(profile)
            
            # Получаем каналы и команды
            login_success = client.login()
            
            if request.method == "POST":
                action = request.POST.get("action")
                if action == "send_message":
                    channel_id = request.POST.get("channel_id")
                    message = request.POST.get("message", "").strip()
                    if channel_id and message:
                        result = client.send_message(channel_id, message)
                        if not result:
                            error_message = "Не удалось отправить сообщение"
                elif action == "create_dm" and login_success:
                    # Обработка создания нового прямого канала
                    create_dm_user_id = request.POST.get("create_dm_user_id")
                    if create_dm_user_id:
                        new_channel = client.create_direct_channel(create_dm_user_id)
                        if new_channel:
                            return redirect(f"?channel={new_channel.get('id')}")
                        else:
                            error_message = "Не удалось создать чат с выбранным пользователем"
            if login_success:
                all_channels = client.get_channels()
                teams = client.get_teams()
                
                # Разделяем каналы на обычные и прямые (DM)
                regular_channels = [ch for ch in all_channels if ch.get("type") != "D"]
                # Получаем количество непрочитанных сообщений и последнее сообщение для обычных каналов
                for channel in regular_channels:
                    unread_count = client.get_channel_unread_count(channel.get("id", ""))
                    channel["unread_count"] = unread_count
                    # Получаем последнее сообщение для сортировки
                    posts = client.get_channel_posts(channel.get("id", ""), limit=1)
                    if posts and len(posts) > 0:
                        last_post = posts[0]
                        create_at = last_post.get("create_at")
                        if hasattr(create_at, "timestamp"):
                            channel["last_post_time"] = int(create_at.timestamp() * 1000)
                        elif isinstance(create_at, (int, float)):
                            channel["last_post_time"] = int(create_at)
                        else:
                            channel["last_post_time"] = 0
                    else:
                        channel["last_post_time"] = 0
                
                # Сортируем обычные каналы по времени последнего сообщения (новые сверху)
                regular_channels.sort(key=lambda x: x.get("last_post_time", 0), reverse=True)
                
                # Получаем только прямые каналы с сообщениями (уже отсортированы внутри метода)
                direct_channels = client.get_direct_channels(only_with_messages=True)
                
                # Получаем всех пользователей для создания нового чата
                all_users = client.get_all_users(per_page=200)
                # Исключаем текущего пользователя
                all_users = [u for u in all_users if u.get("id") != client.user_id]
                
                # Объединяем для общего списка
                channels = regular_channels + direct_channels
                
                # Получаем выбранный канал
                channel_id = request.GET.get("channel")
                if channel_id:
                    # Находим канал в списке
                    for ch in channels:
                        if ch.get("id") == channel_id:
                            selected_channel = ch
                            break
                    
                    if selected_channel:
                        channel_posts = client.get_channel_posts(channel_id)
                        channel_members = client.get_channel_members(channel_id)
                        
                        # Обрабатываем файлы в сообщениях
                        for post in channel_posts:
                            if post.get("files"):
                                for file_info in post.get("files", []):
                                    file_id = file_info.get("id")
                                    if file_id:
                                        # Создаем локальный путь для файла
                                        file_name = file_info.get("name", f"file_{file_id}")
                                        # Очищаем имя файла от недопустимых символов
                                        import re
                                        file_name = re.sub(r'[^\w\-_\.]', '_', file_name)
                                        local_path = os.path.join(
                                            settings.MATTERMOST_FILES_ROOT,
                                            file_id[:2],  # Поддиректория по первым 2 символам ID
                                            file_name
                                        )
                                        
                                        # Скачиваем файл, если его еще нет
                                        if not os.path.exists(local_path):
                                            client.download_file(file_id, local_path)
                                        
                                        # Добавляем локальный путь к информации о файле
                                        if os.path.exists(local_path):
                                            file_info["local_path"] = local_path
                                            file_info["url"] = f"/mattermost-files/{file_id[:2]}/{file_name}"
            else:
                # Более детальное сообщение об ошибке
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Mattermost login failed for user {request.user.username}. Username: {profile.mattermost_username[:3] if profile.mattermost_username else 'None'}..., Server: {mm_settings.server_url}, SSL verify: {mm_settings.verify_ssl}")
                
                if mm_settings.verify_ssl:
                    error_message = (
                        "Не удалось подключиться к серверу Mattermost из-за ошибки SSL сертификата. "
                        "Варианты решения:\n"
                        "1. Добавьте сертификат сервера в папку certs/ и пересоберите Docker образ\n"
                        "2. Отключите проверку SSL в настройках Mattermost (раздел 'Настройки портала' → 'Сервер Mattermost')"
                    )
                else:
                    error_message = "Не удалось подключиться к серверу Mattermost. Проверьте логин и пароль, а также настройки сервера в разделе 'Настройки портала'."
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Mattermost error: {str(e)}", exc_info=True)
            error_message = f"Ошибка при работе с Mattermost: {str(e)}"

    return render(
        request,
        "organization/chats.html",
        {
            "channels": channels,
            "regular_channels": regular_channels if login_success else [],
            "direct_channels": direct_channels if login_success else [],
            "all_users": all_users if login_success else [],
            "teams": teams if login_success else [],
            "selected_channel": selected_channel,
            "channel_posts": channel_posts,
            "channel_members": channel_members,
            "error_message": error_message,
            "mattermost_configured": mattermost_configured,
            "user_configured": user_configured,
        },
    )


def about(request):
    """
    Страница "О системе" с информацией о корпоративном портале.
    """
    from .models import PortalSettings
    
    portal_settings = PortalSettings.get_solo()
    
    return render(
        request,
        "organization/about.html",
        {
            "portal_settings": portal_settings,
        },
    )


@login_required
@require_http_methods(["POST"])
def save_dashboard_order(request):
    """
    Сохраняет порядок блоков на ленте пользователя.
    """
    import json
    from accounts.models import UserProfile
    
    try:
        data = json.loads(request.body)
        blocks_order = data.get('blocks_order', [])
        
        if not isinstance(blocks_order, list):
            return JsonResponse({'error': 'blocks_order должен быть массивом'}, status=400)
        
        # Проверяем существование колонки через SQL
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'accounts_userprofile' 
                AND column_name = 'dashboard_blocks_order'
            """)
            column_exists = cursor.fetchone() is not None
        
        if not column_exists:
            return JsonResponse({'error': 'Миграция для dashboard_blocks_order еще не применена. Пожалуйста, примените миграции.'}, status=503)
        
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        profile.dashboard_blocks_order = blocks_order
        profile.save()
        
        return JsonResponse({'success': True, 'blocks_order': blocks_order})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат JSON'}, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error saving dashboard order: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def vacations(request):
    """
    Отображение таблицы отпусков всех сотрудников с фильтрацией и экспортом.
    Теперь также позволяет создавать отпуск для текущего пользователя.
    """
    from django.db.models import Q
    from django.contrib import messages
    from datetime import date, datetime
    from accounts.models import UserProfile
    
    error_message = None
    
    # Обработка создания отпуска
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "vacation":
            profile, _ = UserProfile.objects.get_or_create(user=request.user)
            
            starts = request.POST.getlist("start_date")
            ends = request.POST.getlist("end_date")
            types = request.POST.getlist("vacation_type")
            
            # Подсчитываем дни для новых периодов
            new_periods_days = 0
            new_periods = []
            
            for start_str, end_str, vtype in zip(starts, ends, types):
                if not start_str or not end_str:
                    continue
                try:
                    start = datetime.strptime(start_str, "%Y-%m-%d").date()
                    end = datetime.strptime(end_str, "%Y-%m-%d").date()
                    
                    if start > end:
                        error_message = "Дата начала не может быть позже даты окончания"
                        break
                    
                    # Подсчитываем дни для этого периода
                    delta = end - start
                    period_days = delta.days + 1
                    new_periods_days += period_days
                    new_periods.append((start, end, vtype or "main"))
                except ValueError:
                    error_message = "Неверный формат даты"
                    break
            
            if not error_message and new_periods:
                # Определяем год для проверки (используем год первого периода)
                check_year = new_periods[0][0].year
                
                # Получаем уже существующие дни отпуска за этот год
                existing_days = VacationPeriod.get_user_vacation_days_for_year(
                    request.user, check_year
                )
                
                # Проверяем общее количество дней
                total_days = existing_days + new_periods_days
                
                if total_days > 28:
                    error_message = (
                        f"Невозможно внести отпуск: превышение лимита отпускных дней в году. "
                        f"Уже использовано: {existing_days} дней, добавляется: {new_periods_days} дней, "
                        f"итого: {total_days} дней. Максимально допустимо: 28 дней в году."
                    )
                else:
                    # Создаем заявку
                    vacation_request = VacationRequest.objects.create(
                        user=request.user,
                        department=profile.department,
                        unit=profile.unit,
                        position=profile.position,
                        last_name=profile.last_name or request.user.last_name or "",
                        first_name=profile.first_name or request.user.first_name or "",
                        middle_name=profile.middle_name or "",
                    )
                    
                    # Создаем периоды
                    for start, end, vtype in new_periods:
                        VacationPeriod.objects.create(
                            request=vacation_request,
                            start_date=start,
                            end_date=end,
                            vacation_type=vtype,
                        )
                    
                    messages.success(request, "Отпуск успешно добавлен в график")
                    return redirect("vacations")
            elif not error_message:
                error_message = "Необходимо указать хотя бы один период отпуска"
            
            if error_message:
                messages.error(request, error_message)
    
    # Получаем все периоды отпусков с информацией о заявках
    periods = VacationPeriod.objects.select_related('request', 'request__department', 'request__unit', 'request__position').all().order_by('start_date')
    
    # Фильтрация
    search_query = request.GET.get('search', '').strip()
    department_filter = request.GET.get('department', '').strip()
    unit_filter = request.GET.get('unit', '').strip()
    
    if search_query:
        periods = periods.filter(
            Q(request__last_name__icontains=search_query) |
            Q(request__first_name__icontains=search_query) |
            Q(request__middle_name__icontains=search_query)
        )
    
    if department_filter:
        periods = periods.filter(request__department__id=department_filter)
    
    if unit_filter:
        periods = periods.filter(request__unit__id=unit_filter)
    
    # Получаем уникальные значения для фильтров
    all_departments = Department.objects.all().order_by('name')
    all_units = Unit.objects.all().order_by('name')
    
    # Формируем данные для таблицы
    vacations_data = []
    for period in periods:
        request_obj = period.request
        full_name = f"{request_obj.last_name} {request_obj.first_name} {request_obj.middle_name}".strip()
        vacations_data.append({
            'full_name': full_name,
            'last_name': request_obj.last_name,
            'first_name': request_obj.first_name,
            'middle_name': request_obj.middle_name,
            'start_date': period.start_date,
            'end_date': period.end_date,
            'department': request_obj.department.name if request_obj.department else '',
            'unit': request_obj.unit.name if request_obj.unit else '',
            'position': request_obj.position.name if request_obj.position else '',
            'vacation_type': period.get_vacation_type_display(),
        })
    
    # Экспорт
    export_format = request.GET.get('export')
    if export_format == 'xls':
        return export_vacations_xls(vacations_data)
    elif export_format == 'pdf':
        return export_vacations_pdf(vacations_data)
    
    # Получаем информацию о текущем пользователе для формы
    user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    current_year = date.today().year
    user_vacation_days = VacationPeriod.get_user_vacation_days_for_year(
        request.user, current_year
    )
    
    return render(
        request,
        "organization/vacations.html",
        {
            "vacations": vacations_data,
            "all_departments": all_departments,
            "all_units": all_units,
            "search_query": search_query,
            "department_filter": department_filter,
            "unit_filter": unit_filter,
            "user_profile": user_profile,
            "user_vacation_days": user_vacation_days,
            "current_year": current_year,
        },
    )


def export_vacations_xls(vacations_data):
    """Экспорт отпусков в Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "График отпусков"
    
    # Заголовки
    headers = ["ФИО", "Дата начала", "Дата окончания", "Отдел", "Департамент", "Должность", "Тип отпуска"]
    ws.append(headers)
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    for vacation in vacations_data:
        ws.append([
            vacation['full_name'],
            vacation['start_date'].strftime('%d.%m.%Y'),
            vacation['end_date'].strftime('%d.%m.%Y'),
            vacation['unit'],
            vacation['department'],
            vacation.get('position', ''),
            vacation['vacation_type'],
        ])
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="vacations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def export_vacations_pdf(vacations_data):
    """Экспорт отпусков в PDF."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="vacations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("График отпусков сотрудников", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Подготовка данных для таблицы
    data = [["ФИО", "Дата начала", "Дата окончания", "Отдел", "Департамент", "Должность", "Тип отпуска"]]
    
    for vacation in vacations_data:
        data.append([
            vacation['full_name'],
            vacation['start_date'].strftime('%d.%m.%Y'),
            vacation['end_date'].strftime('%d.%m.%Y'),
            vacation['unit'],
            vacation['department'],
            vacation.get('position', ''),
            vacation['vacation_type'],
        ])
    
    # Создание таблицы
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def mattermost_file_view(request, file_subdir: str, file_name: str):
    """
    Предоставляет доступ к файлам из Mattermost.
    """
    from .mattermost_client import MattermostClient
    from .models import MattermostSettings
    from accounts.models import UserProfile
    
    try:
        profile = request.user.profile
        mm_settings = MattermostSettings.get_solo()
        
        if not profile.mattermost_username or not profile.mattermost_password:
            raise Http404("Mattermost не настроен")
        
        # Путь к файлу
        file_path = os.path.join(
            str(settings.MATTERMOST_FILES_ROOT),
            file_subdir,
            file_name
        )
        
        # Проверяем, что файл существует и находится в правильной директории
        if not os.path.exists(file_path) or not file_path.startswith(str(settings.MATTERMOST_FILES_ROOT)):
            raise Http404("Файл не найден")
        
        # Отдаем файл
        return FileResponse(
            open(file_path, 'rb'),
            filename=file_name,
            as_attachment=False
        )
    except Exception:
        raise Http404("Файл не найден")


def documents(request):
    from documents.views import document_list

    return document_list(request)


def processes(request):
    return render(
        request,
        "section_placeholder.html",
        {
            "title": "Бизнес-процессы",
            "subtitle": "Настройка и контроль бизнес-процессов: согласование заявок, отпусков и др.",
        },
    )


def hr_services(request):
    return render(
        request,
        "section_placeholder.html",
        {
            "title": "HR-сервисы",
            "subtitle": "Кадровые сервисы: адаптация сотрудников, учёт рабочего времени, отчёты.",
        },
    )


@login_required
@user_passes_test(_is_portal_admin)
def portal_settings_page(request):
    """
    Панель администрирования портала.
    Доступна только администраторам/суперпользователям.
    Позволяет настраивать:
    - интерфейс и дизайн (PortalSettings)
    - почтовый сервер (MailServerSettings)
    - формировать ссылки‑приглашения
    - просматривать базовую информацию о пользователях
    """

    portal_settings = PortalSettings.get_solo()
    mail_settings = MailServerSettings.get_solo()
    caldav_settings = CaldavSettings.get_solo()
    mattermost_settings = MattermostSettings.get_solo()

    portal_form = PortalSettingsForm(request.POST or None, request.FILES or None, instance=portal_settings, prefix="portal")
    mail_form = MailServerSettingsForm(request.POST or None, instance=mail_settings, prefix="mail")
    caldav_form = CaldavSettingsForm(request.POST or None, instance=caldav_settings, prefix="caldav")
    mattermost_form = MattermostSettingsForm(request.POST or None, instance=mattermost_settings, prefix="mattermost")
    invitation_form = InvitationCreateForm(request.POST or None, prefix="invite")

    if request.method == "POST":
        form_type = request.POST.get("form_type")

        if form_type == "portal" and portal_form.is_valid():
            portal_form.save()
        elif form_type == "mail" and mail_form.is_valid():
            mail_form.save()
        elif form_type == "caldav" and caldav_form.is_valid():
            caldav_form.save()
        elif form_type == "mattermost" and mattermost_form.is_valid():
            mattermost_form.save()
        elif form_type == "invite" and invitation_form.is_valid():
            invitation: Invitation = invitation_form.save(commit=False)
            invitation.created_by = request.user
            invitation.expires_at = timezone.now() + timedelta(days=7)
            invitation.save()
        elif form_type == "add_department_head":
            # Добавление нового руководителя департамента из списка сотрудников
            user_id = request.POST.get("user_id")
            department_name = request.POST.get("department_name", "")
            if user_id and department_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    department_head, created = DepartmentHead.objects.get_or_create(
                        user=user,
                        department_name=department_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        department_head.last_name = profile.last_name or ""
                        department_head.first_name = profile.first_name or ""
                        department_head.middle_name = profile.middle_name or ""
                        department_head.department = str(profile.department) if profile.department else ""
                        department_head.unit = str(profile.unit) if profile.unit else ""
                        department_head.position = str(profile.position) if profile.position else ""
                        department_head.phone_personal = profile.phone_personal or ""
                        department_head.phone_internal = profile.phone_internal or ""
                        department_head.photo = profile.photo
                    department_head.save()
                except User.DoesNotExist:
                    pass
        elif form_type == "remove_department_head":
            # Удаление руководителя департамента
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    DepartmentHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass
        elif form_type == "add_unit_head":
            # Добавление нового руководителя отдела из списка сотрудников
            user_id = request.POST.get("user_id")
            unit_name = request.POST.get("unit_name", "")
            if user_id and unit_name:
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    unit_head, created = UnitHead.objects.get_or_create(
                        user=user,
                        unit_name=unit_name,
                        defaults={}
                    )
                    # Обновляем данные из UserProfile
                    if hasattr(user, "profile"):
                        profile = user.profile
                        unit_head.last_name = profile.last_name or ""
                        unit_head.first_name = profile.first_name or ""
                        unit_head.middle_name = profile.middle_name or ""
                        unit_head.department = str(profile.department) if profile.department else ""
                        unit_head.unit = str(profile.unit) if profile.unit else ""
                        unit_head.position = str(profile.position) if profile.position else ""
                        unit_head.phone_personal = profile.phone_personal or ""
                        unit_head.phone_internal = profile.phone_internal or ""
                        unit_head.photo = profile.photo
                    unit_head.save()
                except User.DoesNotExist:
                    pass
        elif form_type == "remove_unit_head":
            # Удаление руководителя отдела
            head_id = request.POST.get("head_id")
            if head_id:
                try:
                    UnitHead.objects.filter(pk=head_id).delete()
                except Exception:
                    pass

    invitations = Invitation.objects.order_by("-created_at")[:20]
    users = User.objects.order_by("username")[:20]
    
    # Получаем всех активных сотрудников для выбора
    all_employees = User.objects.filter(is_active=True).select_related('profile').order_by('username')
    
    # Получаем списки руководителей
    department_heads = DepartmentHead.objects.select_related('user').order_by('department_name', 'last_name', 'first_name')
    unit_heads = UnitHead.objects.select_related('user').order_by('unit_name', 'last_name', 'first_name')
    
    # Получаем списки департаментов и отделов для выбора при добавлении руководителя
    all_departments = Department.objects.all().order_by('name')
    all_units = Unit.objects.all().order_by('name')

    return render(
        request,
        "admin/portal_settings.html",
        {
            "portal_form": portal_form,
            "mail_form": mail_form,
            "caldav_form": caldav_form,
            "mattermost_form": mattermost_form,
            "invitation_form": invitation_form,
            "invitations": invitations,
            "users": users,
            "all_employees": all_employees,
            "department_heads": department_heads,
            "unit_heads": unit_heads,
            "all_departments": all_departments,
            "all_units": all_units,
        },
    )


def user_roles(request):
    return render(
        request,
        "section_placeholder.html",
        {
            "title": "Пользователи и роли",
            "subtitle": "Управление пользователями портала, ролями и правами доступа.",
        },
    )
